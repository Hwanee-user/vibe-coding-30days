"""Day03 - Parking Violation Counter

Read daily violation counts from Excel and print:
- monthly totals
- weekday totals (Mon~Sun)
- weekday vs weekend totals

Usage:
  python day03_violation_counter.py data.xlsx
  python day03_violation_counter.py data.xlsx --sheet "Sheet1"
"""

from __future__ import annotations

import argparse
import datetime as dt
from collections import defaultdict
from pathlib import Path
from typing import Dict, Optional, Tuple
from openpyxl import load_workbook


WEEKDAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def _column_to_index(col: Optional[str], default: int) -> int:
    """Convert Excel column letter (A, B, ...) to 1-based index."""
    if col is None:
        return default
    s = col.strip()
    if not s:
        return default
    if s.isdigit():
        return int(s)
    n = 0
    for ch in s.upper():
        if not ("A" <= ch <= "Z"):
            continue
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n if n > 0 else default


def _parse_excel_date(value: object) -> Optional[dt.date]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value

    if isinstance(value, (int, float)):
        excel_origin = dt.datetime(1899, 12, 30)
        try:
            return (excel_origin + dt.timedelta(days=int(value))).date()
        except (OSError, OverflowError, ValueError):
            return None

    text = str(value).strip().replace(".", "-")
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%m-%d-%Y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    try:
        return dt.date.fromisoformat(text)
    except ValueError:
        return None


def _parse_count(value: object) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _find_columns(header_row) -> Dict[str, int]:
    """Find column index from header row by Korean / English labels."""
    idx = {"date": None, "count": None}
    for col, cell in enumerate(header_row, start=1):
        if cell is None:
            continue
        name = str(cell.value).strip().lower()
        if idx["date"] is None and any(k in name for k in ("일자", "날짜", "date")):
            idx["date"] = col
        if idx["count"] is None and any(k in name for k in ("위반", "건수", "count", "합계")):
            idx["count"] = col
    return idx


def analyze(filepath: str, sheet_name: Optional[str], header_row: int, date_col: Optional[str], count_col: Optional[str]):
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    if header_row < 1:
        raise ValueError("header-row must be 1 or larger")

    dcol = _column_to_index(date_col, 1)
    ccol = _column_to_index(count_col, 2)

    if date_col is None or count_col is None:
        auto_cols = _find_columns(ws[header_row])
        if date_col is None and auto_cols["date"] is not None:
            dcol = auto_cols["date"]
        if count_col is None and auto_cols["count"] is not None:
            ccol = auto_cols["count"]

    monthly: Dict[str, int] = defaultdict(int)
    weekday: Dict[int, int] = defaultdict(int)
    weekday_sum = 0
    weekend_sum = 0
    total = 0
    skipped = 0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        date_value = row[dcol - 1] if dcol - 1 < len(row) else None
        count_value = row[ccol - 1] if ccol - 1 < len(row) else None

        d = _parse_excel_date(date_value)
        if d is None:
            skipped += 1
            continue

        cnt = _parse_count(count_value)
        if cnt is None:
            skipped += 1
            continue

        monthly[f"{d.year:04d}-{d.month:02d}"] += cnt
        widx = d.weekday()
        weekday[widx] += cnt
        if widx >= 5:
            weekend_sum += cnt
        else:
            weekday_sum += cnt
        total += cnt

    print("Parking Violation Summary")
    print(f"Total violation count: {total:,}")
    print(f"Skipped rows: {skipped}")
    print("")

    print("Monthly totals:")
    for month_key in sorted(monthly):
        print(f"- {month_key}: {monthly[month_key]:,}")
    print("")

    print("Weekday totals:")
    for i, name in enumerate(WEEKDAY_NAMES):
        print(f"- {name}: {weekday.get(i, 0):,}")
    print("")

    print(f"Weekday total (Mon-Fri): {weekday_sum:,}")
    print(f"Weekend total (Sat-Sun): {weekend_sum:,}")


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Count parking violations by month and weekday")
    p.add_argument("excel", type=str, help="path to excel file (.xlsx)")
    p.add_argument("--sheet", type=str, default=None, help="sheet name (optional)")
    p.add_argument("--header-row", type=int, default=1, help="header row number (default: 1)")
    p.add_argument("--date-col", type=str, default=None, help="date column letter or number (optional)")
    p.add_argument("--count-col", type=str, default=None, help="count column letter or number (optional)")
    return p


def main():
    args = _build_parser().parse_args()
    path = Path(args.excel)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    analyze(
        filepath=str(path),
        sheet_name=args.sheet,
        header_row=args.header_row,
        date_col=args.date_col,
        count_col=args.count_col,
    )


if __name__ == "__main__":
    main()
